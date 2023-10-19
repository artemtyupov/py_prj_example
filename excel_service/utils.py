import logging
import os
import shutil
from pathlib import Path
from typing import Dict, List, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.worksheet import Worksheet
from resources import (OTHER_SPENDINGS_TEMPLATE_FILENAME,
                       REAL_COSTS_TEMPLATE_FILENAME,
                       REPORT_TEMPLATE_FILENAME)

from shared_files.models import CClientSpending, CRealPriceItem


async def create_overview_list(wb: Workbook):
    list1: Worksheet = wb.create_sheet("Overview Бизнеса")
    list1.cell(row=1, column=1).value = ""

    # column A
    list1.cell(row=2, column=1).value = "Чистая прибыль"
    list1.cell(row=3, column=1).value = "% Маржинальность"
    list1.cell(row=4, column=1).value = "Продажи, руб"
    list1.cell(row=5, column=1).value = "% Выкупа"
    list1.cell(row=6, column=1).value = "Возвраты, руб"
    list1.cell(row=7, column=1).value = "% Возвратов"
    list1.cell(row=8, column=1).value = "Логистика, руб"
    list1.cell(row=9, column=1).value = "Комиссия WB"
    list1.cell(row=10, column=1).value = "% Комиссии"
    list1.cell(row=11, column=1).value = "Хранение, руб"
    list1.cell(row=12, column=1).value = "Себестоимость, руб"
    list1.cell(row=13, column=1).value = "Штрафы, руб"
    list1.cell(row=14, column=1).value = "Компенсации WB, руб"
    list1.cell(row=15, column=1).value = "Траты вне ВБ, руб"
    
    # column B
    list1.cell(row=1, column=2).value = "Неделя 1"
    list1.cell(row=1, column=3).value = "Неделя 2"
    list1.cell(row=1, column=4).value = "Неделя 3"
    list1.cell(row=1, column=5).value = "Неделя 4"
    list1.cell(row=1, column=6).value = "Неделя 5"

async def create_data_list(wb: Workbook):
    pass

async def create_report():
    main_dir = os.path.dirname(__file__)
    filepath = os.path.join(main_dir, f"templates/{REPORT_TEMPLATE_FILENAME}")
    
    workbook: Workbook = load_workbook(filepath)
    create_data_list(workbook)

async def validate_excel_data(user_client_id, type, new_file_data):
    main_dir = os.path.dirname(__file__)
    templates_dir = os.path.join(main_dir, "templates")
    filepath = os.path.join(templates_dir, f"tmp_xl_{user_client_id}.xlsx")
    with open(filepath, 'wb') as f:
        f.write(new_file_data)
        
    workbook: Workbook = load_workbook(filepath)
    sheet: Worksheet = workbook.active
    if type == 'other_spendings':
        is_valid = True
        for i in range(2, sheet.max_row + 1):
            spending_name = sheet[f'B{i}'].value
            frequncy = sheet[f'C{i}'].value
            price = sheet[f'D{i}'].value
            
            nulls_num = 0
            if spending_name == "" or spending_name == None:
                nulls_num += 1
                
            if frequncy not in ['Ежедневно', 'Еженедельно', 'Ежемесячно', 'На каждый товар', 'Разовая трата'] or frequncy == None:
                nulls_num += 1
                
            if price == 0 or price == None:
                nulls_num += 1
                
            if nulls_num != 0 and nulls_num != 3:
                is_valid = False

        workbook.close()
        os.remove(filepath)
        
        return is_valid
    else:
        for i in range(2, sheet.max_row + 1):
            item_real_price = sheet[f'C{i}'].value
            if item_real_price == None or item_real_price == "" or item_real_price == 0:
                if sheet[f'A{i}'].value != None and sheet[f'A{i}'].value != "" and sheet[f'A{i}'].value!= 0:
                    return False
                

        workbook.close()
        os.remove(filepath)
        
        return True
    
async def get_updated_excel_data(user_client_id: int, type: str, new_file_data, old_models: Union[List[CClientSpending], List[CRealPriceItem]] = None) -> Union[List[CClientSpending], List[CRealPriceItem]]:
    main_dir = os.path.dirname(__file__)
    templates_dir = os.path.join(main_dir, "templates")
    filepath = os.path.join(templates_dir, f"tmp_xl_{user_client_id}.xlsx")
    with open(filepath, 'wb') as f:
        f.write(new_file_data)
        
    workbook: Workbook = load_workbook(filepath)
    sheet: Worksheet = workbook.active
    retval_models = []
    if type == 'other_spendings':
        models_map: Dict[int, CClientSpending] = {}
        if old_models:
            for model in old_models:
                models_map[model.spending_id] = model
            
        for i in range(2, sheet.max_row + 1):
            id = sheet[f'A{i}'].value
            spending_name = sheet[f'B{i}'].value
            frequncy = sheet[f'C{i}'].value
            price = sheet[f'D{i}'].value
            
            # для новых строчек выставляем минусовой id, чтобы потом их отличить
            if id == None:
                retval_models.append(CClientSpending(spending_name=spending_name, 
                                                     user_client_id=user_client_id,
                                                     frequency=frequncy, 
                                                     price=price,
                                                     status=CClientSpending.Status.NEW.value))
                
            else:
                if id not in models_map: # TODO ошибку прокидывать
                    logging.error("BUG : Появился id, которого раньше не было")
                    raise Exception("BUG : Появился id, которого раньше не было")
                
                old_model = models_map[id]
                old_spending_name = old_model.spending_name
                old_price = old_model.price
                old_frequncy = old_model.frequency
                
                if old_spending_name == None or old_price == None or old_frequncy == None:
                    logging.error("BUG : Пустое поле в old_models")
                    continue
                
                if old_spending_name == spending_name and old_price == price and old_frequncy == frequncy:
                    continue
                
                retval_models.append(CClientSpending(spending_name=spending_name, 
                                                     user_client_id=user_client_id,
                                                     frequency=frequncy, 
                                                     price=price,
                                                     status=CClientSpending.Status.OLD_UPDATED.value))
            
        workbook.close()
        os.remove(filepath)
        
        for old_id in models_map.keys():
            if old_id not in [model.spending_id for model in retval_models]:
                retval_models.append(CClientSpending(spending_name=None, 
                                                     user_client_id=user_client_id,
                                                     frequency=None, 
                                                     price=None,
                                                     status=CClientSpending.Status.OLD_REMOVED.value))
    else:
        for i in range(2, sheet.max_row + 1):
            item_id = sheet[f'A{i}'].value
            item_real_price = sheet[f'C{i}'].value
            if item_id != None and item_id !=  "" and item_id != 0:
                retval_models.append(CRealPriceItem(item_id=item_id, 
                                                    user_client_id=user_client_id,
                                                    item_real_price=item_real_price))
    return retval_models
    
#TODO удаление строк из excel багает
async def get_excel_data(user_client_id: int, type: str, old_models: Union[List[CClientSpending], List[CRealPriceItem]] = None):
    filename = ""
    if type == 'other_spendings':
        filename = OTHER_SPENDINGS_TEMPLATE_FILENAME
    else:
        filename = REAL_COSTS_TEMPLATE_FILENAME
    
    main_dir = os.path.dirname(__file__)
    templates_dir = os.path.join(main_dir, "templates")
    filepath = os.path.join(templates_dir, filename)
    if old_models == None or len(old_models) == 0:
        if filename == REAL_COSTS_TEMPLATE_FILENAME:
            logging.error("Real costs table can not be null, beacause we must get items from DB")
            raise Exception("Real costs table can not be null, beacause we must get items from DB")
        
        return filepath

    file_name = Path(filepath).stem
    tmp_filename = file_name + f"_{user_client_id}.xlsx"
    tmp_filepath = os.path.join(templates_dir, tmp_filename)
    shutil.copyfile(filepath, tmp_filepath)
    workbook: Workbook = load_workbook(tmp_filepath)
    sheet: Worksheet = workbook.active
    if old_models:
        if type == 'other_spendings':
            i = 2
            for model in old_models:
                sheet[f'A{i}'] = model.spending_id
                sheet[f'B{i}'] = model.spending_name
                sheet[f'C{i}'] = model.frequency
                sheet[f'D{i}'] = model.price
                i += 1
        else:
            i = 2
            for model in old_models:
                sheet[f'A{i}'] = model.item_id
                sheet[f'B{i}'] = model.vendor_code
                sheet[f'C{i}'] = model.item_real_price
                i += 1

    for row in sheet.rows:
        if row.index == 1:
            continue
        
        for cell in row:
            cell.protection = Protection(locked=False)
    
    os.remove(tmp_filepath)
    workbook.save(tmp_filepath)
    workbook.close()
    
    return tmp_filepath